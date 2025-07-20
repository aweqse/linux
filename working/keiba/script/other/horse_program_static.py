#coding: UTF-8
#このプログラムは金曜日の深夜および土曜日の深夜にタスクスケジューラー等で自動実行することを前提にしている
import datetime
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
import mysql.connector
import subprocess
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import shutil
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import re

options = webdriver.ChromeOptions()
options.add_argument("--headless=new")
options.add_argument('--disable-gpu')
options.add_argument('--ignore-certificate-errors')
options.add_argument('--allow-running-insecure-content')
options.add_argument('--disable-web-security')
options.add_argument('--blink-settings=imagesEnabled=false')
options.add_argument('--ignore-certificate-errors')
driver = webdriver.Chrome(options=options)

conn = mysql.connector.connect(
host='localhost',
port='3306',
user='root',
password='Fujieda1217', 
database='keiba'
)
cursor = conn.cursor()

# sleeptoolを終了させる
#cmd = 'taskkill /im Sleeptool.exe /f'
#returncode = subprocess.call(cmd)

# リストを初期化する
race_url_kantou = []
race_url_kansai = []
race_url_tihou = []
all_url_race = []

# 現在の日付と曜日を取得する
dt_now = datetime.datetime.now()
year = dt_now.year

# ダイアログに情報を入力する
print("取得したいレースの日付を入力してください")
month = input("月:")
day = input("日:")

#手動で馬柱を生成する場合
#month=5
#day=11

# 文字列に変換する
year = str(year)
month = str(month)
day = str(day)

# 桁数が一桁の場合に"0"を挿入する
judgment_day = len(day)
if judgment_day == 1:
    day = "0"+day

 # 桁数が一桁の場合に"0"を挿入する
judgment_month = len(month)
if judgment_month == 1:
    month = "0"+month

#エクセルファイルを生成する。
#Desktopとデスクトップの違いがあるので注意
#excel_path="C:\\Users\\aweqs\\OneDrive\\プログラム\\馬柱生成プログラム\\出力先\\"+year+month+day+"馬柱.xlsx"
#desktop="C:\\Users\\aweqs\\Desktop\\"+year+month+day+"馬柱.xlsx"
excel_path="C:\workspace\\プログラム\\馬柱生成プログラム\\出力先\\"+year+month+day+"馬柱.xlsx"
desktop="C:\\Users\\aweqs\\Desktop\\"+year+month+day+"馬柱.xlsx"

wb = openpyxl.Workbook()
wb.save(desktop)

race_url_list=[]
access_index=0

# urlにアクセスする
url = "https://race.netkeiba.com/top/race_list.html?kaisai_date="+year+month+day
driver.get(url)
WebDriverWait(driver, 15).until(EC.presence_of_all_elements_located)

#一日単位のレースのURLを取得する
elements=driver.find_elements(By.XPATH,("/html/body/div[1]/div/div[1]/div/div[2]/div/div[1]//a[@href]"))
for element in elements:                
    element=element.get_attribute("href")
    WebDriverWait(driver, 15).until(EC.presence_of_all_elements_located)
    race_url_list.append(element)

#余分なリンク先を削除する
day_index=0
max_day_index=len(race_url_list)
while day_index<max_day_index:
    if ("payback" in race_url_list[day_index])==True or ("javascript" in race_url_list[day_index])==True  or ("top" in race_url_list[day_index])==True or ("movie" in race_url_list[day_index])==True or ("tv" in race_url_list[day_index])==True:
        del race_url_list[day_index]
        max_day_index=len(race_url_list)
    else:
        day_index=day_index+1

#URLがレース結果の場合は出馬表に変換する
check_3=0
while len(race_url_list)>check_3:
    check_str=race_url_list[check_3]
    if ("result" in race_url_list[check_3])==True:
        check_str=re.search(r"\d+",check_str)
        check_str=check_str.group()
        race_url_list[check_3]="https://race.netkeiba.com/race/shutuba.html?race_id="+check_str+"&rf=race_submenu"
        check_3=check_3+1
    check_3=check_3+1

print("URL取得完了")    
sleep(3)

xpath_count=1
race_cache=len(race_url_list)
umabashira_list=[]
head_list=[]

while race_cache>access_index:#ここは一括で取得したURLを条件とする
    #本番では以下３つをコメントアウトしない
    driver.get(race_url_list[access_index])
    
    #以下テスト
    #driver.get("https://race.netkeiba.com/race/shutuba.html?race_id=202408030411")
    #access_index=access_index+100

    WebDriverWait(driver, 15).until(EC.presence_of_all_elements_located)
    hantei_4=driver.find_element(By.XPATH,"/html/body/div[1]/div[2]/div/div[1]/div[3]/div[1]/span")
    hantei_4=hantei_4.text

    horse_name_list=[]

    #ヘッダーをを取得する
    hantei_2=driver.find_element(By.XPATH,("/html/body/div[1]/div[2]/div/div[1]/div[3]/div[2]"))
    hantei_2=hantei_2.text.split()

    #牝馬限定戦の判定をする
    #hinba_indexはレース名に”牝”がつくものを除外するためとhantei_2[9]なので処理軽減のために５から始める
    hinba_flag=0
    hinba_index=5
    while hinba_index<len(hantei_2):    
        if ("牝" in hantei_2[hinba_index])==True:
            if ("牡" in hantei_2[hinba_index])==False:
                hinba_flag=1
            del hantei_2[hinba_index]
            break
        hinba_index=hinba_index+1

    #配列修正　不要な情報を削除する
    del_index=0
    len_hantei_2=len(hantei_2)
    while del_index<len_hantei_2:
        if ("(春)" in hantei_2[del_index])==True or ("(秋)" in hantei_2[del_index])==True:
            del_index=del_index+1
            continue
        elif ("発走" in hantei_2[del_index])==True or ("馬場" in hantei_2[del_index])==True or ("/" in hantei_2[del_index])==True or ("天候" in hantei_2[del_index])==True or ("本賞金" in hantei_2[del_index])==True or ("(" in hantei_2[del_index])==True or (")" in hantei_2[del_index])==True or ("[" in hantei_2[del_index])==True or ("]" in hantei_2[del_index])==True or ("見習"in hantei_2[del_index])==True or ("九州産馬"in hantei_2[del_index])==True:
            del hantei_2[del_index]
            len_hantei_2=len(hantei_2)
            continue
        del_index=del_index+1

    #配列修正　取りやめの場合挿入する
    if ("取り止め" in hantei_2[0])==True:
        hantei_2.insert(4,"0回")
    
    #配列修正　障害戦でダートと芝両方の馬場状態が取得されるので削除する
    check_1=hantei_2[2]
    check_2=hantei_2[3]
    if ("芝" in check_1) or ("ダート" in check_1)==True: 
        if ("芝" in check_2) or ("ダート" in check_2)==True:
            del hantei_2[3]
            del hantei_2[4]

    #配列修正　アルファベット（右・Bなど）が含まれている場合削除する
    check_8=hantei_2[3]
    check_9=hantei_2[4]
    if ("A" in check_8) or ("B" in check_8) or ("C" in check_8) or ("D" in check_8)==True:
        del hantei_2[3]
    if ("A" in check_9) or ("B" in check_9) or ("C" in check_9) or ("D" in check_9)==True:
        del hantei_2[4]
        
    #外回りの判定をする
    #hinba_indexはレース名に”外”がつくものを除外するため1から始める
    out_flag=0
    out_index=1
    while out_index<len(hantei_2):    
        if ("外" in hantei_2[out_index])==True:
            del hantei_2[out_index]
            out_flag=1
            break
        out_index=out_index+1
    
    #長距離の場合の（内2週）を削除する
    str_search=0
    while len(hantei_2)>str_search:
        if ("内2周" in hantei_2[str_search])==True:
            del hantei_2[str_search]
            str_search=str_search+1
        str_search=str_search+1

    number_of_race=hantei_4.replace("R","")
    course=hantei_2[1]
    kaisaijou=hantei_2[3]

    #同じコースで内・外の区別がある場合に判定する
    if kaisaijou=="京都" and (course=="芝1400m" or course=="芝1600m"):
        if out_flag==1:
            out_in="外"
            course=course+"外"
        elif out_flag==0:
            out_in="内"
            course=course+"内"
    elif kaisaijou=="新潟" and course=="芝2000m":
        if out_flag==1:
            out_in="外"
            course=course+"外"
        elif out_flag==0:
            out_in="内"
            course=course+"内"
    else:
        out_in="NaN"

    head_count=hantei_2[8].replace("頭", "")

    head_count=int(head_count)

    while xpath_count<=head_count:
        #枠、馬番、馬名を取得する
        xpath_count=str(xpath_count)
        element=driver.find_elements(By.XPATH,("/html/body/div[1]/div[3]/div[2]/table/tbody/tr["+xpath_count+"]"))        
        if len(element)==0:
            element=driver.find_elements(By.XPATH,("/html/body/div[1]/div[3]/div[2]/div[1]/table/tbody/tr["+xpath_count+"]"))        
        element=element[0].text.split()           

        #調教師と美浦・栗東を加工する
        if ("美浦" in element[7])==True:
            tilyoukilyoushi=element[7].replace("美浦","")
            element.insert(8,tilyoukilyoushi)
            element.insert(7,"美浦")
            del element[8]
        elif ("栗東" in element[7])==True:
            tilyoukilyoushi=element[7].replace("栗東","")
            element.insert(8,tilyoukilyoushi)
            element.insert(7,"栗東")
            del element[8]
        else:
            tilyoukilyoushi=element[7].replace("海外","")
            element.insert(8,tilyoukilyoushi)
            element.insert(7,"海外")    
            del element[8]      

        del element[9]
        del element[2]
        del_count=len(element)
        while del_count>8:
            del element[8]
            del_count=len(element)
        umabashira_list.append(element)
        xpath_count=int(xpath_count)
        xpath_count=xpath_count+1

        horse_name_list.append(element[2])
        print("出馬情報取得完了!!")
    
    #MySQLから出走履歴を読み込む
    horse_name_index=0
    horse_name_index=0
    horse_cache=""
    horse_query=""
    while len(horse_name_list)>horse_name_index:
        horse_chace=horse_name_list[horse_name_index]
        if horse_name_index==len(horse_name_list)-1:
            horse_cache="horse_name='"+horse_name_list[horse_name_index]+"';"
        else:
            horse_cache="horse_name='"+horse_name_list[horse_name_index]+"' or "
        horse_query=horse_query+horse_cache
        horse_name_index=horse_name_index+1

    df_query="select wakuban,umaban,horse_name,race_time,juni,year_,mouth,day,kaisaijou,cource,in_out,baba,race_rank,race_age,old,kinryou,jockey from all_race where "+horse_query
    df = pd.read_sql(df_query,conn)
    
    print("MySQLからの読み込み完了")
    
    #新しいシートを追加する
    startrow_count=24
    sheet_name = kaisaijou + hantei_4
    #wb.create_sheet(sheet_name)

    #Mysqlの内容をエクセルに書き込む
    with pd.ExcelWriter(desktop, engine="openpyxl",mode='a') as writer:
        df.to_excel(writer, sheet_name=sheet_name, startrow=startrow_count,startcol=1,index=False)

    #フィルターを追加する
    wb = openpyxl.load_workbook(desktop)
    ws = wb[sheet_name]
    ws.auto_filter.ref = "A25:U25"
    wb.save(desktop)

    print("MySQLのデータの書き込み完了")
    
    #dfはmysqlのselectで使うので馬柱は下記のように手動で書き込む
    #対象レースの情報を書き込む
    writer_book = openpyxl.load_workbook(desktop, read_only=False)
    sleep(2)
    sheet=writer_book[kaisaijou+hantei_4]
    sheet["B1"]="レース名"
    sheet["C1"]="コース"
    sheet["D1"]="回数"
    sheet["E1"]="競馬場"
    sheet["F1"]="日目"
    sheet["G1"]="条件"
    sheet["H1"]="クラス"
    sheet["I1"]="斤量"
    sheet["J1"]="頭数"
    sheet["K1"]="牝馬限定"

    sheet["B2"]=hantei_2[0]
    sheet["C2"]=course
    sheet["D2"]=hantei_2[2] 
    sheet["E2"]=hantei_2[3] 
    sheet["F2"]=hantei_2[4] 
    sheet["G2"]=hantei_2[5] 
    sheet["H2"]=hantei_2[6] 
    sheet["I2"]=hantei_2[7] 
    sheet["J2"]=hantei_2[8] 
    if hinba_flag==0:
        sheet["K2"]="NaN"
    elif hinba_flag==1:
        sheet["K2"]="牝"

    sheet["B4"]="枠番"
    sheet["C4"]="馬番"
    sheet["D4"]="馬名"
    sheet["E4"]="馬齢"
    sheet["F4"]="斤量"  
    sheet["G4"]="騎手"
    sheet["H4"]="東西"
    sheet["I4"]="厩舎"

    excel_index=5
    umabashira_list_index=0
    while len(umabashira_list)>umabashira_list_index:
        excel_index=str(excel_index)

        sheet_B="B"+excel_index
        sheet_C="C"+excel_index
        sheet_D="D"+excel_index
        sheet_E="E"+excel_index
        sheet_F="F"+excel_index
        sheet_G="G"+excel_index
        sheet_H="H"+excel_index
        sheet_I="I"+excel_index

        sheet[sheet_B]=umabashira_list[umabashira_list_index][0]
        sheet[sheet_C]=umabashira_list[umabashira_list_index][1]
        sheet[sheet_D]=umabashira_list[umabashira_list_index][2]
        sheet[sheet_E]=umabashira_list[umabashira_list_index][3]
        sheet[sheet_F]=umabashira_list[umabashira_list_index][4]
        sheet[sheet_G]=umabashira_list[umabashira_list_index][5]
        sheet[sheet_H]=umabashira_list[umabashira_list_index][6]
        sheet[sheet_I]=umabashira_list[umabashira_list_index][7]

        excel_index=int(excel_index)
        excel_index=excel_index+1
        umabashira_list_index=umabashira_list_index+1
    
    writer_book.save(desktop)
    
    print(sheet_name+"の書き込み完了")

    horse_name_list=[]
    umabashira_list=[]
    access_index=access_index+1
    xpath_count=1
sleep(3)
writer_book.close()
shutil.move(desktop,excel_path)
print("すべての処理が完了!!")


#/html/body/div[1]/div[3]/div[2]/table/tbody/tr[1]
#/html/body/div[1]/div[3]/div[2]/table/tbody/tr[13]
#/html/body/div[1]/div[3]/div[2]/table/tbody/tr[1]
#/html/body/div[1]/div[3]/div[2]/table/tbody/tr[10]

#/html/body/div[1]/div[2]/div/div[1]/div[3]/div[2]/div[3]/span[8]
#/html/body/div[1]/div[2]/div/div[1]/div[3]/div[2]/div[3]/span[8]